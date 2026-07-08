import os
import json
import base64
from flask import Flask, request, jsonify
from flask_cors import CORS
import anthropic
from openpyxl import Workbook
import PyPDF2
import fitz
import io
from datetime import datetime

app = Flask(__name__)
CORS(app, origins="*")
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB

client = anthropic.Anthropic(api_key=os.environ.get("ANTHROPIC_API_KEY"))

# ─────────────────────────────────────────────
# UTILITIES
# ─────────────────────────────────────────────

def extract_text_from_pdf(pdf_file):
    pdf_bytes = pdf_file.read()
    pdf_file.seek(0)
    reader = PyPDF2.PdfReader(io.BytesIO(pdf_bytes))
    text = ""
    for page in reader.pages:
        text += (page.extract_text() or "") + "\n"
    if len(text.strip()) > 100:
        return text, None
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    images = []
    # Rate tables and policies are always in the first few pages.
    # Cap at 8 pages to keep the upload under Railway's proxy size limit.
    MAX_PAGES = 8
    for i, page in enumerate(doc):
        if i >= MAX_PAGES:
            break
        pix = page.get_pixmap(matrix=fitz.Matrix(1.5, 1.5))
        img_bytes = pix.tobytes("jpeg", jpg_quality=85)
        images.append(base64.standard_b64encode(img_bytes).decode("utf-8"))
    doc.close()
    return None, images


def build_claude_message(text, images, instruction):
    if text:
        return [{"type": "text", "text": f"{instruction}\n\n{text}"}]
    content = [{"type": "image", "source": {"type": "base64", "media_type": "image/jpeg", "data": img}} for img in images]
    content.append({"type": "text", "text": instruction})
    return content


def clean_json_response(text):
    text = text.strip()
    if "```" in text:
        for part in text.split("```"):
            part = part.strip().lstrip("json").strip()
            if part.startswith("{"):
                return part
    return text


def call_claude(system_prompt, user_message):
    """
    Call Claude using streaming. For large contracts the JSON can still
    exceed token limits, so we use a two-pass approach:
    Pass 1: extract everything EXCEPT room_seasons
    Pass 2: extract room_seasons only
    Then merge.
    """
    full_text = ""
    with client.beta.messages.stream(
        model="claude-sonnet-4-6",
        max_tokens=16000,
        system=system_prompt,
        messages=[{"role": "user", "content": user_message}],
        betas=["output-128k-2025-02-19"],
    ) as stream:
        for text in stream.text_stream:
            full_text += text

    raw = clean_json_response(full_text)
    try:
        return json.loads(raw)
    except json.JSONDecodeError:
        raise


CONTRACT_HEADER_PROMPT = """
You are a hotel rate sheet expert. Read this hotel contract and return ONLY this JSON — no room rates, no markdown:
{
  "hotel_name": "string",
  "reservation_date_from": "DD/MM/YYYY",
  "reservation_date_till": "DD/MM/YYYY",
  "meal_plans": ["Bed and Breakfast"],
  "supplement_rules": {
    "hb_per_adult": 0,
    "fb_per_adult": 0,
    "extra_bed_adult": 0,
    "extra_bed_child_under_6": 0,
    "extra_bed_child_6_to_12": 0,
    "child_meal_hb_under_6": 0,
    "child_meal_hb_6_to_12": 0,
    "child_meal_fb_under_6": 0,
    "child_meal_fb_6_to_12": 0
  },
  "occupancy_policy": {
    "max_adults": 2,
    "max_extra_beds": 1,
    "child_age_brackets": [
      {
        "label": "CHILD",
        "min_age": 0,
        "max_age": 11.99,
        "free_sharing": true,
        "paid_sharing": false,
        "free_extra_bed": false,
        "paid_extra_bed": false,
        "supplement_key": "under_6"
      }
    ]
  }
}
Rules:
- meal_plans: include HB/FB only if explicitly offered
- supplement_rules: set 0 if free/complimentary
- occupancy_policy: read child policy carefully — one bracket per distinct age group
- All dates DD/MM/YYYY
- Output raw JSON only
"""

CONTRACT_SEASONS_PROMPT = """
You are a hotel rate extraction expert. Read this hotel contract and return ONLY a JSON array of room seasons — no other text:
[
  {
    "room": "string",
    "base_bb": 0,
    "season_begin": "DD/MM/YYYY",
    "season_end": "DD/MM/YYYY",
    "season_type": "High",
    "res_date_from": "DD/MM/YYYY",
    "res_date_till": "DD/MM/YYYY"
  }
]
Rules:
- base_bb = Double Occupancy (2A) BB rate per room per night
- If the rate table has occupancy columns (1A, 2A, 2A+1C, 3A), use the 2A column
- One entry per room type per season date range
- If a season has non-consecutive dates, create a separate entry per date range
- season_type: exactly "Low", "Shoulder", "High", or "Peak" (map Festive/F to "Peak")
- res_date_from = contract start date — same for every entry
- res_date_till = contract end date — same for every entry, NEVER the season_end date
- All dates DD/MM/YYYY
- Output raw JSON array only — no object wrapper, no markdown
"""

PROMOTION_HEADER_PROMPT = """
You are a hotel rate sheet expert. Read this promotion/SPO and return ONLY this JSON — no room rates, no markdown:
{
  "hotel_name": "string",
  "promo_code": "string",
  "reservation_date_from": "DD/MM/YYYY",
  "reservation_date_till": "DD/MM/YYYY",
  "mlos": 1,
  "mlos_till": 366,
  "meal_plans": ["Bed and Breakfast"],
  "supplement_rules": {
    "hb_per_adult": 0,
    "fb_per_adult": 0,
    "extra_bed_adult": 0,
    "extra_bed_child_under_6": 0,
    "extra_bed_child_6_to_12": 0,
    "child_meal_hb_under_6": 0,
    "child_meal_hb_6_to_12": 0,
    "child_meal_fb_under_6": 0,
    "child_meal_fb_6_to_12": 0
  },
  "occupancy_policy": {
    "max_adults": 2,
    "max_extra_beds": 1,
    "child_age_brackets": [
      {
        "label": "CHILD",
        "min_age": 0,
        "max_age": 11.99,
        "free_sharing": true,
        "paid_sharing": false,
        "free_extra_bed": false,
        "paid_extra_bed": false,
        "supplement_key": "under_6"
      }
    ]
  }
}
Rules:
- reservation_date_from/till = the booking window (not the stay dates)
- mlos = minimum nights (default 1), mlos_till = maximum nights (default 366)
- supplement_rules: use contract context if not in promotion, set 0 if free
- occupancy_policy: read child policy carefully
- All dates DD/MM/YYYY
- Output raw JSON only
"""

PROMOTION_SEASONS_PROMPT = """
You are a hotel rate extraction expert. Read this promotion and return ONLY a JSON array of room seasons — no other text:
[
  {
    "room": "string",
    "base_rate": 0,
    "meal_plan": "Bed and Breakfast",
    "season_begin": "DD/MM/YYYY",
    "season_end": "DD/MM/YYYY",
    "season_type": "Low",
    "res_date_from": "DD/MM/YYYY",
    "res_date_till": "DD/MM/YYYY"
  }
]
Rules:
- base_rate = the FINAL discounted promo rate — never the contracted rate
- One entry per room type per date row in the promotion table
- meal_plan = the meal plan the rate is quoted at
- season_type: exactly "Low", "Shoulder", "High", or "Peak" (Festive/F = "Peak")
- res_date_from/till = the promotion booking window (same for all entries)
- All dates DD/MM/YYYY
- Output raw JSON array only — no object wrapper, no markdown
"""


def stream_claude(system_prompt, user_message, max_tokens=8000):
    """Single streaming Claude call, returns parsed text."""
    full_text = ""
    with client.beta.messages.stream(
        model="claude-sonnet-4-6",
        max_tokens=max_tokens,
        system=system_prompt,
        messages=[{"role": "user", "content": user_message}],
        betas=["output-128k-2025-02-19"],
    ) as stream:
        for text in stream.text_stream:
            full_text += text
    return clean_json_response(full_text)


def call_claude_split(system_prompt, user_message, is_promotion=False):
    """
    Two dedicated Claude calls — header info first, then room seasons.
    Each call has its own clean focused system prompt.
    """
    if is_promotion:
        header_raw = stream_claude(PROMOTION_HEADER_PROMPT, user_message, max_tokens=4000)
        seasons_raw = stream_claude(PROMOTION_SEASONS_PROMPT, user_message, max_tokens=16000)
    else:
        header_raw = stream_claude(CONTRACT_HEADER_PROMPT, user_message, max_tokens=4000)
        seasons_raw = stream_claude(CONTRACT_SEASONS_PROMPT, user_message, max_tokens=16000)

    header_data = json.loads(header_raw)

    seasons_clean = seasons_raw.strip()
    if seasons_clean.startswith("["):
        seasons = json.loads(seasons_clean)
    else:
        wrapped = json.loads(seasons_clean)
        seasons = wrapped.get("room_seasons", list(wrapped.values())[0] if wrapped else [])

    header_data["room_seasons"] = seasons
    return header_data


def parse_date(date_str):
    try:
        return datetime.strptime(date_str.strip(), "%d/%m/%Y").date()
    except:
        return None


def validate_dates(data):
    contract_end = data.get("reservation_date_till", "")
    contract_start = data.get("reservation_date_from", "")
    for season in data.get("room_seasons", []):
        if not season.get("res_date_from"):
            season["res_date_from"] = contract_start
        if not season.get("res_date_till") or season.get("res_date_till") == season.get("season_end"):
            season["res_date_till"] = contract_end
    return data


DATE_COLUMNS = {
    "Season begin", "Season end",
    "Reservation date from", "Reservation date till",
    "Check-in from", "Check-in till", "Check-out from", "Check-out till",
}


def generate_excel_from_data(rows, headers):
    import datetime as dt
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    date_col_indices = {i for i, h in enumerate(headers) if h in DATE_COLUMNS}
    for row in rows:
        ws.append(row)
    for col_idx in date_col_indices:
        col_letter = ws.cell(row=1, column=col_idx + 1).column_letter
        for cell in ws[col_letter][1:]:
            if isinstance(cell.value, dt.date):
                cell.number_format = "DD/MM/YYYY"
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return base64.b64encode(buffer.read()).decode("utf-8")


# ─────────────────────────────────────────────
# DYNAMIC OCCUPANCY GENERATION
# ─────────────────────────────────────────────

def generate_occupancy_combinations(policy):
    max_adults = policy.get("max_adults", 2)
    brackets = policy.get("child_age_brackets", [])
    max_extra_beds = policy.get("max_extra_beds", 1)

    combos = []

    free_sharing_brackets = [b for b in brackets if b.get("free_sharing")]
    paid_sharing_brackets = [b for b in brackets if b.get("paid_sharing")]
    paid_extra_brackets   = [b for b in brackets if b.get("paid_extra_bed")]

    # Adults only
    for adults in range(1, max_adults + 1):
        extra_adults = max(0, adults - 2)
        base_adults = min(adults, 2)
        if extra_adults > max_extra_beds:
            continue
        combos.append({
            "label": f"{adults}ADL",
            "adults": base_adults, "adult_extra_beds": extra_adults,
            "child_free_sharing": 0, "child_paid_sharing_under_6": 0,
            "child_paid_sharing_6_to_12": 0, "child_free_extra": 0,
            "child_paid_extra_under_6": 0, "child_paid_extra_6_to_12": 0,
        })

    # 2 Adults + free-sharing children
    for b in free_sharing_brackets:
        for n in range(1, 3):
            combos.append({
                "label": f"2ADL+{n}-CHILD SHARING",
                "adults": 2, "adult_extra_beds": 0,
                "child_free_sharing": n, "child_paid_sharing_under_6": 0,
                "child_paid_sharing_6_to_12": 0, "child_free_extra": 0,
                "child_paid_extra_under_6": 0, "child_paid_extra_6_to_12": 0,
            })
        break  # one free-sharing bracket is enough

    # 2 Adults + paid-sharing children
    for b in paid_sharing_brackets:
        age_lbl = b.get("label", "CHILD")
        is_u6 = b.get("max_age", 99) <= 6
        for n in range(1, 3):
            combos.append({
                "label": f"2ADL+{n}-{age_lbl} SHARING",
                "adults": 2, "adult_extra_beds": 0,
                "child_free_sharing": 0,
                "child_paid_sharing_under_6": n if is_u6 else 0,
                "child_paid_sharing_6_to_12": 0 if is_u6 else n,
                "child_free_extra": 0,
                "child_paid_extra_under_6": 0, "child_paid_extra_6_to_12": 0,
            })

    # Mixed free + paid sharing
    if free_sharing_brackets and paid_sharing_brackets:
        pb = paid_sharing_brackets[0]
        is_u6 = pb.get("max_age", 99) <= 6
        combos.append({
            "label": "2ADL+1CH FREE+1CH SHARING",
            "adults": 2, "adult_extra_beds": 0,
            "child_free_sharing": 1,
            "child_paid_sharing_under_6": 1 if is_u6 else 0,
            "child_paid_sharing_6_to_12": 0 if is_u6 else 1,
            "child_free_extra": 0,
            "child_paid_extra_under_6": 0, "child_paid_extra_6_to_12": 0,
        })

    # 2 Adults + paid extra-bed children
    for b in paid_extra_brackets:
        age_lbl = b.get("label", "CHILD")
        is_u6 = b.get("max_age", 99) <= 6
        for n in range(1, max_extra_beds + 1):
            combos.append({
                "label": f"2ADL+{n}-{age_lbl} EXTRA",
                "adults": 2, "adult_extra_beds": 0,
                "child_free_sharing": 0,
                "child_paid_sharing_under_6": 0, "child_paid_sharing_6_to_12": 0,
                "child_free_extra": 0,
                "child_paid_extra_under_6": n if is_u6 else 0,
                "child_paid_extra_6_to_12": 0 if is_u6 else n,
            })

    # Mixed extra bed (two paid brackets, two extra beds)
    if len(paid_extra_brackets) >= 2 and max_extra_beds >= 2:
        combos.append({
            "label": "2ADL+1CH U6+1CH O6 EXTRA",
            "adults": 2, "adult_extra_beds": 0,
            "child_free_sharing": 0,
            "child_paid_sharing_under_6": 0, "child_paid_sharing_6_to_12": 0,
            "child_free_extra": 0,
            "child_paid_extra_under_6": 1, "child_paid_extra_6_to_12": 1,
        })

    # 1 Adult + paid extra-bed child
    for b in paid_extra_brackets:
        age_lbl = b.get("label", "CHILD")
        is_u6 = b.get("max_age", 99) <= 6
        combos.append({
            "label": f"1ADL+1-{age_lbl} EXTRA",
            "adults": 1, "adult_extra_beds": 0,
            "child_free_sharing": 0,
            "child_paid_sharing_under_6": 0, "child_paid_sharing_6_to_12": 0,
            "child_free_extra": 0,
            "child_paid_extra_under_6": 1 if is_u6 else 0,
            "child_paid_extra_6_to_12": 0 if is_u6 else 1,
        })

    # Deduplicate
    seen, unique = set(), []
    for c in combos:
        if c["label"] not in seen:
            seen.add(c["label"])
            unique.append(c)
    return unique


# ─────────────────────────────────────────────
# HEADERS
# ─────────────────────────────────────────────

CONTRACT_HEADERS = [
    "Hotel", "Room", "Accommodation", "Meal",
    "Season begin", "Season end",
    "Reservation date from", "Reservation date till",
    "Nights", "Hotel net price", "Number of markups",
    "Currency (code)", "Currency", "Season type",
    "Market code", "Price type",
    "Staying nights from", "Staying nights till", "Booking code"
]

PROMOTION_HEADERS = [
    "SPO No", "Price type", "Hotel", "Room", "Accommodation", "Meal",
    "Hotel net price", "Currency (code)", "Market code",
    "Season begin", "Season end",
    "Days before check-in from", "Reservation date from", "Reservation date till",
    "Check-in from", "Check-in till", "Check-out from",
    "Staying nights from", "Check-out till", "Nights",
    "Nights from", "Nights till", "Number of markups",
    "Nights free", "Season type", "Days before check-in till",
    "Staying nights till", "Booking code"
]

# ─────────────────────────────────────────────
# CLAUDE PROMPTS
# ─────────────────────────────────────────────

CONTRACT_EXTRACTION_PROMPT = """
You are a hotel rate sheet expert for Voyage Tours, a Dubai-based tour operator.

Extract ALL rate data from this hotel contract and return a single valid JSON object.

Return ONLY this JSON — no markdown, no explanation, no trailing commas:
{
  "hotel_name": "string",
  "reservation_date_from": "DD/MM/YYYY",
  "reservation_date_till": "DD/MM/YYYY",
  "meal_plans": ["Bed and Breakfast"],
  "supplement_rules": {
    "hb_per_adult": 0,
    "fb_per_adult": 0,
    "extra_bed_adult": 0,
    "extra_bed_child_under_6": 0,
    "extra_bed_child_6_to_12": 0,
    "child_meal_hb_under_6": 0,
    "child_meal_hb_6_to_12": 0,
    "child_meal_fb_under_6": 0,
    "child_meal_fb_6_to_12": 0
  },
  "occupancy_policy": {
    "max_adults": 2,
    "max_extra_beds": 1,
    "child_age_brackets": [
      {
        "label": "CHILD",
        "min_age": 0,
        "max_age": 11.99,
        "free_sharing": true,
        "paid_sharing": false,
        "free_extra_bed": false,
        "paid_extra_bed": false,
        "supplement_key": "under_6"
      }
    ]
  },
  "room_seasons": [
    {
      "room": "string",
      "base_bb": 0,
      "season_begin": "DD/MM/YYYY",
      "season_end": "DD/MM/YYYY",
      "season_type": "High",
      "res_date_from": "DD/MM/YYYY",
      "res_date_till": "DD/MM/YYYY"
    }
  ]
}

OCCUPANCY POLICY RULES:
- max_adults: maximum adults the room can hold (2 or 3)
- max_extra_beds: maximum extra beds allowed (usually 1)
- child_age_brackets: one entry per distinct child age group
  - label: short label e.g. "CHILD", "CHILD06", "CHILD12"
  - min_age / max_age: the age range (e.g. 0 to 5.99, 6 to 11.99)
  - free_sharing: true if child shares existing bedding at no charge
  - paid_sharing: true if child shares existing bedding but pays meal supplement
  - free_extra_bed: true if child gets extra bed at no charge
  - paid_extra_bed: true if child needs extra bed and pays supplement
  - supplement_key: "under_6" if max_age <= 6, else "6_to_12"

RATE RULES:
- base_bb = Double Occupancy (2A) BB rate per room per night
- If contract shows rate columns per occupancy (1A, 2A, 2A+1C, 3A), use the 2A column
- One room_seasons entry per room type per season date range
- If a season spans non-consecutive dates, create a separate entry per date range

DATE RULES:
- reservation_date_from = contract start date
- reservation_date_till = contract end date (last valid stay)
- res_date_from per season = same as reservation_date_from
- res_date_till per season = same as reservation_date_till — NEVER the season_end date
- season_begin / season_end = actual stay dates for that period
- All dates in DD/MM/YYYY format

OTHER RULES:
- season_type: exactly "Low", "Shoulder", "High", or "Peak"
- Add HB/FB to meal_plans only if explicitly offered
- Set supplement to 0 if free or complimentary
- Output valid JSON only — no trailing commas, no comments, no markdown
"""

PROMOTION_EXTRACTION_PROMPT = """
You are a hotel rate sheet expert for Voyage Tours, a Dubai-based tour operator.

Extract ALL rate data from this hotel promotion/SPO and return a single valid JSON object.

CRITICAL: Use the FINAL DISCOUNTED RATE as base_rate — never the contracted rate.

Return ONLY this JSON — no markdown, no explanation, no trailing commas:
{
  "hotel_name": "string",
  "promo_code": "string",
  "reservation_date_from": "DD/MM/YYYY",
  "reservation_date_till": "DD/MM/YYYY",
  "mlos": 1,
  "mlos_till": 366,
  "meal_plans": ["Bed and Breakfast"],
  "supplement_rules": {
    "hb_per_adult": 0,
    "fb_per_adult": 0,
    "extra_bed_adult": 0,
    "extra_bed_child_under_6": 0,
    "extra_bed_child_6_to_12": 0,
    "child_meal_hb_under_6": 0,
    "child_meal_hb_6_to_12": 0,
    "child_meal_fb_under_6": 0,
    "child_meal_fb_6_to_12": 0
  },
  "occupancy_policy": {
    "max_adults": 2,
    "max_extra_beds": 1,
    "child_age_brackets": [
      {
        "label": "CHILD",
        "min_age": 0,
        "max_age": 11.99,
        "free_sharing": true,
        "paid_sharing": false,
        "free_extra_bed": false,
        "paid_extra_bed": false,
        "supplement_key": "under_6"
      }
    ]
  },
  "room_seasons": [
    {
      "room": "string",
      "base_rate": 0,
      "meal_plan": "Bed and Breakfast",
      "season_begin": "DD/MM/YYYY",
      "season_end": "DD/MM/YYYY",
      "season_type": "Low",
      "res_date_from": "DD/MM/YYYY",
      "res_date_till": "DD/MM/YYYY"
    }
  ]
}

RATE RULES:
- base_rate = final selling rate after discount for that room and date range
- If the promo table has many date rows per room, create one room_seasons entry per room per row
- meal_plan = the meal plan the base_rate is quoted at

DATE RULES:
- reservation_date_from = booking open date for this promotion
- reservation_date_till = last date to book (booking cutoff)
- res_date_from per season = same as reservation_date_from
- res_date_till per season = same as reservation_date_till — NOT the season_end
- season_begin / season_end = actual stay dates for that row
- All dates in DD/MM/YYYY format

OTHER RULES:
- mlos = minimum length of stay (default 1)
- mlos_till = maximum length of stay (default 366)
- season_type: exactly "Low", "Shoulder", "High", or "Peak"
- If promotion does not state child policy, use the contract context provided
- Set supplements to 0 if free/complimentary or not stated
- Output valid JSON only — no trailing commas, no comments, no markdown
"""

# ─────────────────────────────────────────────
# PRICE CALCULATION
# ─────────────────────────────────────────────

def calculate_price(base_bb, meal, occ, rules):
    hb = rules.get("hb_per_adult", 0)
    fb = rules.get("fb_per_adult", 0)
    extra_bed_adult = rules.get("extra_bed_adult", 0)
    extra_bed_child_u6 = rules.get("extra_bed_child_under_6", 0)
    extra_bed_child_6 = rules.get("extra_bed_child_6_to_12", 0)
    child_hb_u6 = rules.get("child_meal_hb_under_6", 0)
    child_hb_6 = rules.get("child_meal_hb_6_to_12", 0)
    child_fb_u6 = rules.get("child_meal_fb_under_6", 0)
    child_fb_6 = rules.get("child_meal_fb_6_to_12", 0)

    if meal in ["Half Board", "HB"]:
        adult_meal = hb; child_meal_u6 = child_hb_u6; child_meal_6 = child_hb_6
    elif meal in ["Full Board", "FB"]:
        adult_meal = fb; child_meal_u6 = child_fb_u6; child_meal_6 = child_fb_6
    else:
        adult_meal = 0; child_meal_u6 = 0; child_meal_6 = 0

    price = base_bb
    price += (occ["adults"] + occ["adult_extra_beds"]) * adult_meal
    price += occ["adult_extra_beds"] * extra_bed_adult

    n_u6 = occ.get("child_paid_extra_under_6", 0)
    n_6  = occ.get("child_paid_extra_6_to_12", 0)
    price += n_u6 * (extra_bed_child_u6 + child_meal_u6)
    price += n_6  * (extra_bed_child_6  + child_meal_6)
    if n_6 >= 2:
        price += extra_bed_child_6  # second child extra bed surcharge

    price += occ.get("child_paid_sharing_under_6", 0) * child_meal_u6
    price += occ.get("child_paid_sharing_6_to_12", 0) * child_meal_6
    return round(price)

# ─────────────────────────────────────────────
# ROW EXPANSION
# ─────────────────────────────────────────────

def expand_contract_rates(hotel_name, room_seasons, meal_plans, supplement_rules, occupancy_policy):
    combos = generate_occupancy_combinations(occupancy_policy)
    rows = []
    for season in room_seasons:
        s_begin = parse_date(season["season_begin"])
        s_end   = parse_date(season["season_end"])
        r_from  = parse_date(season["res_date_from"])
        r_till  = parse_date(season["res_date_till"])
        for meal in meal_plans:
            for occ in combos:
                price = calculate_price(season["base_bb"], meal, occ, supplement_rules)
                row = {
                    "Hotel": hotel_name, "Room": season["room"],
                    "Accommodation": occ["label"], "Meal": meal,
                    "Season begin": s_begin, "Season end": s_end,
                    "Reservation date from": r_from, "Reservation date till": r_till,
                    "Nights": 1, "Hotel net price": price,
                    "Number of markups": 1, "Currency (code)": "AED",
                    "Currency": "Dirham", "Season type": season["season_type"],
                    "Market code": "", "Price type": "Standard",
                    "Staying nights from": 1, "Staying nights till": 366, "Booking code": ""
                }
                rows.append([row.get(h, "") for h in CONTRACT_HEADERS])
    return rows


def expand_promotion_rates(hotel_name, promo_code, room_seasons, meal_plans, supplement_rules, occupancy_policy, mlos, mlos_till):
    combos = generate_occupancy_combinations(occupancy_policy)
    rows = []
    for season in room_seasons:
        s_begin = parse_date(season["season_begin"])
        s_end   = parse_date(season["season_end"])
        r_from  = parse_date(season["res_date_from"])
        r_till  = parse_date(season["res_date_till"])
        for meal in meal_plans:
            for occ in combos:
                price = calculate_price(season["base_rate"], meal, occ, supplement_rules)
                row = {
                    "SPO No": "", "Price type": "Standard",
                    "Hotel": hotel_name, "Room": season["room"],
                    "Accommodation": occ["label"], "Meal": meal,
                    "Hotel net price": price, "Currency (code)": "AED", "Market code": "",
                    "Season begin": s_begin, "Season end": s_end,
                    "Days before check-in from": "",
                    "Reservation date from": r_from, "Reservation date till": r_till,
                    "Check-in from": "", "Check-in till": "", "Check-out from": "",
                    "Staying nights from": mlos, "Check-out till": "", "Nights": 1,
                    "Nights from": "", "Nights till": "", "Number of markups": 1,
                    "Nights free": "", "Season type": season["season_type"],
                    "Days before check-in till": "",
                    "Staying nights till": mlos_till, "Booking code": promo_code
                }
                rows.append([row.get(h, "") for h in PROMOTION_HEADERS])
    return rows

# ─────────────────────────────────────────────
# API ROUTES
# ─────────────────────────────────────────────

@app.route("/api/process", methods=["POST"])
def process_pdfs():
    raw_contract = None
    raw_promotion = None
    try:
        # Accept JSON with base64-encoded PDFs to bypass Railway proxy size limit,
        # or multipart/form-data for smaller files.
        if request.content_type and "application/json" in request.content_type:
            body = request.get_json(force=True)
            if not body or not body.get("contract"):
                return jsonify({"error": "Contract PDF is required"}), 400
            contract_file = io.BytesIO(base64.b64decode(body["contract"]))
            promotion_file = io.BytesIO(base64.b64decode(body["promotion"])) if body.get("promotion") else None
        else:
            contract_file = request.files.get("contract")
            promotion_file = request.files.get("promotion")
            if not contract_file:
                return jsonify({"error": "Contract PDF is required"}), 400

        contract_text, contract_images = extract_text_from_pdf(contract_file)

        contract_data = call_claude_split(
            CONTRACT_EXTRACTION_PROMPT,
            build_claude_message(contract_text, contract_images, "Extract all rate data from this hotel contract:")
        )
        contract_data = validate_dates(contract_data)

        contract_rows = expand_contract_rates(
            hotel_name=contract_data["hotel_name"],
            room_seasons=contract_data["room_seasons"],
            meal_plans=contract_data["meal_plans"],
            supplement_rules=contract_data["supplement_rules"],
            occupancy_policy=contract_data["occupancy_policy"]
        )

        result = {"contract_excel": generate_excel_from_data(contract_rows, CONTRACT_HEADERS)}

        if promotion_file:
            promotion_text, promotion_images = extract_text_from_pdf(promotion_file)
            context_note = ("\n\nCONTRACT CONTEXT (use for supplement rules and child policy if not in promotion):\n" + contract_text) if contract_text else ""

            promotion_data = call_claude_split(
                PROMOTION_EXTRACTION_PROMPT,
                build_claude_message(promotion_text, promotion_images,
                    f"Extract all rate data from this promotion PDF.{context_note}"),
                is_promotion=True
            )
            promotion_data = validate_dates(promotion_data)

            promotion_rows = expand_promotion_rates(
                hotel_name=promotion_data["hotel_name"],
                promo_code=promotion_data.get("promo_code", ""),
                room_seasons=promotion_data["room_seasons"],
                meal_plans=promotion_data["meal_plans"],
                supplement_rules=promotion_data["supplement_rules"],
                occupancy_policy=promotion_data["occupancy_policy"],
                mlos=promotion_data.get("mlos", 1),
                mlos_till=promotion_data.get("mlos_till", 366)
            )
            result["promotion_excel"] = generate_excel_from_data(promotion_rows, PROMOTION_HEADERS)

        return jsonify(result)

    except json.JSONDecodeError as e:
        raw = raw_contract or raw_promotion or "unavailable"
        return jsonify({"error": f"Failed to parse Claude response: {str(e)}", "raw_response": raw[:2000] if isinstance(raw, str) else str(raw)[:2000]}), 500
    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "trace": traceback.format_exc()[-1500:]}), 500


@app.route("/health", methods=["GET"])
def health():
    return jsonify({"status": "ok"})


if __name__ == "__main__":
    app.run(debug=True)
